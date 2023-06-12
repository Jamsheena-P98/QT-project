from openpyxl import load_workbook


def readData(file, sheetName, rowNum, columnNum):
  wb = load_workbook(file)
  ws = wb[sheetName]
  return ws.cell(row=rowNum, column=columnNum).value